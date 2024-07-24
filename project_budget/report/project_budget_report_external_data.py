from odoo import fields, models
from openpyxl import load_workbook
from io import BytesIO
import base64
import datetime


class ReportExternalData(models.Model):
    _name = 'project_budget.report_external_data'
    _description = 'Report External Data'
    _check_company_auto = True
    _inherit = ['mail.thread']

    SUBJECT_STRING = '[Отчет для УК]'  # Тема письма "[Отчет для УК] Стэп Лоджик_23.07.2024"

    company_id = fields.Many2one('res.company', string='Company', required=True)
    report_date = fields.Date(string='Report Date', default=fields.date.today())
    data = fields.Text(string='Data')
    file = fields.Binary(string='File')

    def name_get(self):
        return [(record.id, f"{record.report_date}") for record in self]

    def file_to_data(self, file):
        wb = load_workbook(file, read_only=True, data_only=True)
        ws = wb.active
        data = []
        for record in ws.iter_rows(min_row=6, max_row=6, min_col=4, max_col=None, values_only=True):
            data = str(record).strip('()')
        if data:
            return data

    def message_new(self, msg_dict, custom_values=None):
        subject = msg_dict.get('subject', '')
        attachments = msg_dict.get('attachments', '')
        if self.SUBJECT_STRING in subject and attachments:
            for attachment in attachments:
                company_name, date = subject.split(self.SUBJECT_STRING)[1].split('_')
                data = self.file_to_data(BytesIO(attachment[1]))
                if data:
                    defaults = {
                        'company_id': self.env['res.company'].search([('name', '=', company_name.strip())]).id,
                        'data': data,
                        'file': base64.b64encode(attachment[1]),
                        'report_date': datetime.datetime.strptime(date.strip(), '%d.%m.%Y')
                    }
                    defaults.update(custom_values or {})
                    res = super(ReportExternalData, self).message_new(msg_dict, custom_values=defaults)
                    return res
